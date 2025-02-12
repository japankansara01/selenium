import openpyxl

file = "D:/japan_office/feb/selenium/session 14/Sample.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]

rows = sheet.max_row
cols = sheet.max_column

# Reading all rows and colums from excel sheet
for r in range(1,rows+1):
    for c in range(1, cols+1):
        print(sheet.cell(r,c).value,end='       ')
    print()