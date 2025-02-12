import openpyxl

file = "D:/japan_office/feb/selenium/session 14/Sample1.xlsx"
workbook = openpyxl.load_workbook(file)
# sheet = workbook["Sheet1"]
sheet = workbook.active #Get active sheet from excel

# To Write same data
# for r in range(1, 6):
#     for c in range(1, 4):
#         sheet.cell(r,c).value="Hello"
# workbook.save(file)

sheet.cell(1,1).value=123
sheet.cell(1,2).value="ABC"
sheet.cell(1,3).value="engineer"

sheet.cell(2,1).value=234
sheet.cell(2,2).value="BCD"
sheet.cell(2,3).value="manager"

sheet.cell(3,1).value=345
sheet.cell(3,2).value="CDE"
sheet.cell(3,3).value="QA"

workbook.save(file)